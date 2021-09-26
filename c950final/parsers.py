import openpyxl
import csv


DIST_TABLE = 'c950final/data/WGUPS_Distance_Table.xlsx'
LOC_FILE = 'c950final/data/locations.csv'
EDGE_FILE = 'c950final/data/edges.csv'


def clean(string):
    return str(string).split("\n")[0].strip()


def parse_distance_chart():
    with openpyxl.load_workbook(DIST_TABLE) as wb:
        sheet = wb.active # Only one sheet anyway

        # Parse locations
        locations = {}
        location_id = 1
        for row in range(2, sheet.max_row):
            name = clean(sheet[row][0].value)
            address = clean(sheet[row][1].value)
            locations[name] = {"id": location_id, "name": name, "adr": address}
            location_id += 1

        # Parse chart as graph nodes and edges
        edges = []
        for row in range(2, sheet.max_row):
            source = clean(sheet[row][0].value)
            
            for col in range(2, row):
                distance = sheet[row][col].value
                destination = clean(sheet[1][col].value)
                
                if distance == 0 or distance == None:
                    break # Skip self or empty cells.
                else:
                    edges.append((locations[source]['id'], distance, locations[destination]['id']))

    # Dump locations to file
    with open(LOC_FILE, 'w', newline='') as loc_file:
        writer = csv.writer(loc_file, quoting=csv.QUOTE_NONNUMERIC)
        for loc in sorted(locations.values(), key=lambda item: item['id']): # Write csv sorted by id.
            writer.writerow([loc['id'], loc['name'], loc['adr']])

    # Dump edges to file
    with open(EDGE_FILE, 'w') as edge_file:
        for edge in edges:
            edge_file.write("{},{},{}\n".format(edge[0],edge[1],edge[2]))

    print('Done parsing distance chart.')


if __name__ == "__main__":
    parse_distance_chart()
